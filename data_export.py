import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Ruta del archivo CSV
archivo_csv = 'atenciones_transfomados.csv'

try:
    # Leer el archivo CSV
    data = pd.read_csv(archivo_csv)
    
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Atenciones"

    # Obtener fecha y hora actuales
    now = datetime.now()
    fecha_actual = now.strftime("%d/%m/%Y")
    hora_actual = now.strftime("%H:%M:%S")
    nro_registros = len(data)

    # Definir estilos generales
    encabezado_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
    encabezado_font = Font(bold=True, color="000000")
    encabezado_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    bold_font = Font(bold=True, size=12)
    alignment_left = Alignment(horizontal="left", vertical="center")

    # Espacio antes de los datos
    espacio_filas = 5  # Cantidad de filas de espacio antes de la tabla
    fila_inicio_encabezado = 1  # Inicio del encabezado
    fila_inicio_tabla = espacio_filas + 1  # Donde comenzarán los datos

    # Agregar encabezado personalizado
    encabezado = [("Fecha:", fecha_actual), ("Hora:", hora_actual), ("Nro Reg:", nro_registros)]
    for i, (key, value) in enumerate(encabezado, start=fila_inicio_encabezado):
        ws[f"A{i}"].value = key
        ws[f"A{i}"].font = bold_font
        ws[f"A{i}"].alignment = alignment_left
        ws[f"A{i}"].border = border

        ws[f"B{i}"].value = value
        ws[f"B{i}"].font = Font(size=12)
        ws[f"B{i}"].alignment = alignment_left
        ws[f"B{i}"].border = border

    # Agregar encabezados de la tabla con estilo y bordes
    for col_num, value in enumerate(data.columns, 1):
        cell = ws.cell(row=fila_inicio_tabla, column=col_num, value=value)
        cell.fill = encabezado_fill
        cell.font = encabezado_font
        cell.alignment = encabezado_alignment
        cell.border = border

    # Agregar los datos con estilos condicionales y bordes
    for row_num, row in enumerate(dataframe_to_rows(data, index=False, header=False), start=fila_inicio_tabla + 1):
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)

            # Pintar de color naranja las filas con ID Servicio 9 o 10
            if data.iloc[row_num - fila_inicio_tabla - 1]["ID Servicio"] in [1, 2]:
                cell.fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")

            # Aplicar bordes a cada celda
            cell.border = border

            # Centrar los textos en las celdas
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar el ancho de las columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Letra de la columna
        for cell in column:
            try:
                if cell.value:  # Calcular la longitud máxima del contenido
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # Guardar el archivo Excel
    archivo_excel = 'Cruce de Información Niño.xlsx'
    wb.save(archivo_excel)
    print(f"Datos exportados exitosamente a {archivo_excel} con encabezado, bordes y estilos.")

except Exception as e:
    print(f"Error al procesar los datos: {e}")
